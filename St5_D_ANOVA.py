"""
St6_D_ANOVA_Complete.py - Complete ANOVA Analysis for Paper
Generates both global ANOVA table and individual effects tables

Author: Pablo Antonio Matamala Carvajal
Date: 2025-01-02

Generates:
1. ANOVA Table (Global) - Model, Linear, Quadratic, Interaction
2. Individual Effects Table - Each term with p-values
3. Summary Statistics Table
4. Exports to LaTeX, Excel, CSV

Uses real values from St6_C analysis.
"""

import os
import sys
import pickle
import pandas as pd
import numpy as np
import scipy.stats as stats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

#%%============================================================================
# CONFIGURATION
#%%============================================================================

# Input file
ANALYSIS_FILE = "EcoData/St5_Pareto/P_efficiency_scientific_analysis.pkl"
OUTPUT_FOLDER = "EcoData/St5_ANOVA/paper_tables"

ALPHA = 0.05
RESPONSE_NAME = "Power-Mass Ratio"
RESPONSE_UNITS = "W/(η²kg)"
MAX_EFFECTS_INDIVIDUAL_TABLE = 15

#%%============================================================================
# LOAD DATA
#%%============================================================================

def load_analysis():
    """Load St6_C analysis results"""
    print("="*70)
    print("COMPLETE ANOVA TABLES FOR PAPER")
    print("="*70)
    
    if not os.path.exists(ANALYSIS_FILE):
        print(f"\n❌ ERROR: Analysis file not found!")
        print(f"   Expected: {ANALYSIS_FILE}")
        print(f"\n💡 Run St6_C_ScientificPareto.py first")
        sys.exit(1)
    
    print(f"\n📂 Loading: {ANALYSIS_FILE}")
    
    with open(ANALYSIS_FILE, 'rb') as f:
        data = pickle.load(f)
    
    all_effects = data['all_effects']
    significant_effects = data['significant_effects']
    model_stats = data['model_statistics']
    
    print(f"✅ Analysis loaded:")
    print(f"   Total effects: {len(all_effects)}")
    print(f"   Significant effects: {len(significant_effects)}")
    print(f"   Model R²: {model_stats['r2']:.4f}")
    
    return all_effects, significant_effects, model_stats

#%%============================================================================
# CALCULATE ANOVA GLOBAL (BY TYPE)
#%%============================================================================

def calculate_anova_global(all_effects, model_stats):
    """
    Calculate global ANOVA table with real values
    Source: Model, Linear, Quadratic, Interaction, Residual, Total
    """
    
    print(f"\n{'='*50}")
    print("CALCULATING GLOBAL ANOVA TABLE")
    print(f"{'='*50}")
    
    # Separate effects by type
    linear_fx = [e for e in all_effects if e['type'] == 'main']
    quad_fx = [e for e in all_effects if e['type'] == 'quadratic']
    inter_fx = [e for e in all_effects if e['type'] == 'interaction']
    
    n_samples = model_stats['n_samples']
    n_features = model_stats['n_features']
    
    # Degrees of freedom
    df_total = n_samples - 1
    df_model = n_features - 1  # Total terms excluding intercept
    df_residual = n_samples - n_features
    
    df_linear = len(linear_fx)      # 6
    df_quadratic = len(quad_fx)     # 6
    df_interaction = len(inter_fx)  # 15
    
    # Get R² and MSE from model
    R2 = model_stats['r2']
    MSE = model_stats['mse']
    
    # Calculate Sum of Squares
    SS_residual = MSE * df_residual
    SS_total = SS_residual / (1 - R2) if R2 < 1 else SS_residual * 100
    SS_model = SS_total - SS_residual
    
    # Allocate model SS by type based on effect contributions
    # This is based on the magnitude of effects (Type I approximation)
    total_effect_sq = sum([e['effect']**2 for e in all_effects])
    
    if total_effect_sq > 0:
        SS_linear_contrib = sum([e['effect']**2 for e in linear_fx])
        SS_quad_contrib = sum([e['effect']**2 for e in quad_fx])
        SS_inter_contrib = sum([e['effect']**2 for e in inter_fx])
        
        SS_linear = SS_model * (SS_linear_contrib / total_effect_sq)
        SS_quadratic = SS_model * (SS_quad_contrib / total_effect_sq)
        SS_interaction = SS_model * (SS_inter_contrib / total_effect_sq)
    else:
        # Fallback if no effects
        SS_linear = SS_model / 3
        SS_quadratic = SS_model / 3
        SS_interaction = SS_model / 3
    
    # Mean Squares
    MS_model = SS_model / df_model
    MS_residual = SS_residual / df_residual
    MS_linear = SS_linear / df_linear
    MS_quadratic = SS_quadratic / df_quadratic
    MS_interaction = SS_interaction / df_interaction
    
    # F-statistics
    F_model = MS_model / MS_residual
    F_linear = MS_linear / MS_residual
    F_quadratic = MS_quadratic / MS_residual
    F_interaction = MS_interaction / MS_residual
    
    # P-values
    p_model = 1 - stats.f.cdf(F_model, df_model, df_residual)
    p_linear = 1 - stats.f.cdf(F_linear, df_linear, df_residual)
    p_quadratic = 1 - stats.f.cdf(F_quadratic, df_quadratic, df_residual)
    p_interaction = 1 - stats.f.cdf(F_interaction, df_interaction, df_residual)
    
    # Adjusted R²
    R2_adj = 1 - (SS_residual/df_residual) / (SS_total/df_total)
    RMSE = np.sqrt(MSE)
    
    anova_table = {
        'Source': ['Model', '  Linear', '  Quadratic', '  Interaction', 'Residual', 'Total'],
        'SS': [SS_model, SS_linear, SS_quadratic, SS_interaction, SS_residual, SS_total],
        'df': [df_model, df_linear, df_quadratic, df_interaction, df_residual, df_total],
        'MS': [MS_model, MS_linear, MS_quadratic, MS_interaction, MS_residual, ''],
        'F': [F_model, F_linear, F_quadratic, F_interaction, '', ''],
        'p_value': [p_model, p_linear, p_quadratic, p_interaction, '', '']
    }
    
    stats_summary = {
        'R2': R2,
        'R2_adj': R2_adj,
        'RMSE': RMSE,
        'MSE': MSE,
        'n_samples': n_samples
    }
    
    print(f"✅ Global ANOVA calculated:")
    print(f"   Model F({df_model},{df_residual}) = {F_model:.2f}, p = {p_model:.6f}")
    print(f"   Linear F({df_linear},{df_residual}) = {F_linear:.2f}, p = {p_linear:.6f}")
    print(f"   Quadratic F({df_quadratic},{df_residual}) = {F_quadratic:.2f}, p = {p_quadratic:.6f}")
    print(f"   Interaction F({df_interaction},{df_residual}) = {F_interaction:.2f}, p = {p_interaction:.6f}")
    
    return anova_table, stats_summary

#%%============================================================================
# CREATE DATAFRAMES
#%%============================================================================

def create_anova_global_df(anova_table):
    """Create DataFrame for global ANOVA table"""
    
    print(f"\n{'='*50}")
    print("TABLE 1: GLOBAL ANOVA")
    print(f"{'='*50}")
    
    df = pd.DataFrame({
        'Source': anova_table['Source'],
        'SS': anova_table['SS'],
        'df': anova_table['df'],
        'MS': anova_table['MS'],
        'F': anova_table['F'],
        'p-value': anova_table['p_value']
    })
    
    # Format numbers
    def format_value(val, decimals=4):
        if isinstance(val, (int, float)) and val != '':
            if val < 0.001 and val > 0:
                return '<0.001'
            else:
                return f'{val:.{decimals}f}'
        return val
    
    df['SS'] = df['SS'].apply(lambda x: format_value(x, 4))
    df['MS'] = df['MS'].apply(lambda x: format_value(x, 6))
    df['F'] = df['F'].apply(lambda x: format_value(x, 2))
    df['p-value'] = df['p-value'].apply(lambda x: '<0.001' if (isinstance(x, float) and x < 0.001) else format_value(x, 6))
    
    print(f"✅ Global ANOVA table created (6 rows)")
    
    return df

def create_individual_effects_df(significant_effects):
    """Create DataFrame for individual significant effects"""
    
    print(f"\n{'='*50}")
    print("TABLE 2: INDIVIDUAL EFFECTS")
    print(f"{'='*50}")
    
    # Sort by magnitude
    sig_sorted = sorted(significant_effects, key=lambda x: x['effect'], reverse=True)
    top_effects = sig_sorted[:MAX_EFFECTS_INDIVIDUAL_TABLE]
    
    data = []
    for i, effect in enumerate(top_effects, 1):
        # Significance markers
        if effect['p_value'] < 0.001:
            sig = '***'
        elif effect['p_value'] < 0.01:
            sig = '**'
        elif effect['p_value'] < 0.05:
            sig = '*'
        else:
            sig = ''
        
        type_labels = {'main': 'Linear', 'quadratic': 'Quadratic', 'interaction': 'Interaction'}
        
        data.append({
            'Rank': i,
            'Effect': effect['name'],
            'Type': type_labels.get(effect['type'], effect['type']),
            'Coefficient': f"{effect['coefficient']:.6f}",
            'Std. Effect': f"{effect['effect']:.4f}",
            'p-value': '<0.001' if effect['p_value'] < 0.001 else f"{effect['p_value']:.6f}",
            'Sig.': sig
        })
    
    df = pd.DataFrame(data)
    
    print(f"✅ Individual effects table created ({len(top_effects)} effects)")
    
    return df

def create_summary_df(all_effects, significant_effects, stats_summary):
    """Create summary statistics table"""
    
    print(f"\n{'='*50}")
    print("TABLE 3: SUMMARY STATISTICS")
    print(f"{'='*50}")
    
    # Count by type
    n_sig_linear = len([e for e in significant_effects if e['type'] == 'main'])
    n_sig_quad = len([e for e in significant_effects if e['type'] == 'quadratic'])
    n_sig_inter = len([e for e in significant_effects if e['type'] == 'interaction'])
    
    top3 = sorted(significant_effects, key=lambda x: x['effect'], reverse=True)[:3]
    
    data = {
        'Parameter': [
            'Model R²',
            'Adjusted R²',
            'RMSE',
            'MSE',
            'Number of samples',
            '',
            'Total effects tested',
            'Significant effects (p<0.05)',
            '  Linear (significant)',
            '  Quadratic (significant)',
            '  Interaction (significant)',
            '',
            'Most important effect',
            '2nd most important',
            '3rd most important'
        ],
        'Value': [
            f"{stats_summary['R2']:.4f}",
            f"{stats_summary['R2_adj']:.4f}",
            f"{stats_summary['RMSE']:.6f}",
            f"{stats_summary['MSE']:.6f}",
            str(stats_summary['n_samples']),
            '',
            str(len(all_effects)),
            str(len(significant_effects)),
            str(n_sig_linear),
            str(n_sig_quad),
            str(n_sig_inter),
            '',
            f"{top3[0]['name']} ({top3[0]['effect']:.4f})" if len(top3) > 0 else '',
            f"{top3[1]['name']} ({top3[1]['effect']:.4f})" if len(top3) > 1 else '',
            f"{top3[2]['name']} ({top3[2]['effect']:.4f})" if len(top3) > 2 else ''
        ]
    }
    
    df = pd.DataFrame(data)
    
    print(f"✅ Summary table created")
    
    return df

#%%============================================================================
# EXPORT TO LATEX
#%%============================================================================

def export_to_latex(df_anova, df_effects, df_summary, stats_summary):
    """Export all tables to LaTeX"""
    
    print(f"\n{'='*50}")
    print("EXPORTING TO LATEX")
    print(f"{'='*50}")
    
    latex_folder = os.path.join(OUTPUT_FOLDER, "latex")
    os.makedirs(latex_folder, exist_ok=True)
    
    # Table 1: ANOVA Global
    latex_anova = os.path.join(latex_folder, "table1_anova_global.tex")
    with open(latex_anova, 'w', encoding='utf-8') as f:
        f.write("% Table 1: Global ANOVA\n")
        f.write("\\begin{table}[ht]\n")
        f.write("\\centering\n")
        f.write(f"\\caption{{Analysis of Variance for the quadratic response surface model of {RESPONSE_NAME}}}\n")
        f.write("\\label{tab:anova}\n")
        f.write("\\begin{tabular}{lccccc}\n")
        f.write("\\toprule\n")
        f.write("Source & SS & df & MS & \\textit{F} & \\textit{p}-value \\\\\n")
        f.write("\\midrule\n")
        
        for _, row in df_anova.iterrows():
            f.write(f"{row['Source']} & {row['SS']} & {row['df']} & {row['MS']} & {row['F']} & {row['p-value']} \\\\\n")
        
        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\begin{tablenotes}\n")
        f.write("\\small\n")
        f.write(f"\\item \\textit{{Note}}: Model statistics: R² = {stats_summary['R2']:.4f}, ")
        f.write(f"Adjusted R² = {stats_summary['R2_adj']:.4f}, ")
        f.write(f"RMSE = {stats_summary['RMSE']:.4f} {RESPONSE_UNITS}. ")
        f.write("SS = Sum of Squares, df = degrees of freedom, MS = Mean Square.\n")
        f.write("\\end{tablenotes}\n")
        f.write("\\end{table}\n")
    
    print(f"✅ LaTeX ANOVA table: {latex_anova}")
    
    # Table 2: Individual Effects
    latex_effects = os.path.join(latex_folder, "table2_individual_effects.tex")
    with open(latex_effects, 'w', encoding='utf-8') as f:
        f.write("% Table 2: Individual Significant Effects\n")
        f.write("\\begin{table}[ht]\n")
        f.write("\\centering\n")
        f.write(f"\\caption{{Significant standardized effects on {RESPONSE_NAME} (\\textit{{p}}<{ALPHA})}}\n")
        f.write("\\label{tab:effects}\n")
        f.write("\\begin{tabular}{clccccc}\n")
        f.write("\\toprule\n")
        f.write("Rank & Effect & Type & Coefficient & Std. Effect & \\textit{p}-value & Sig. \\\\\n")
        f.write(" & & & & [{RESPONSE_UNITS}] & & \\\\\n")
        f.write("\\midrule\n")
        
        for _, row in df_effects.iterrows():
            f.write(f"{row['Rank']} & ${row['Effect']}$ & {row['Type']} & {row['Coefficient']} & {row['Std. Effect']} & {row['p-value']} & {row['Sig.']} \\\\\n")
        
        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\begin{tablenotes}\n")
        f.write("\\small\n")
        f.write("\\item \\textit{Note}: Standardized effects calculated using RSM methodology. ")
        f.write("Significance: *** \\textit{p}<0.001, ** \\textit{p}<0.01, * \\textit{p}<0.05.\n")
        f.write("\\end{tablenotes}\n")
        f.write("\\end{table}\n")
    
    print(f"✅ LaTeX effects table: {latex_effects}")
    
    # Table 3: Summary
    latex_summary = os.path.join(latex_folder, "table3_summary.tex")
    with open(latex_summary, 'w', encoding='utf-8') as f:
        f.write("% Table 3: Summary Statistics\n")
        f.write("\\begin{table}[ht]\n")
        f.write("\\centering\n")
        f.write("\\caption{Summary of Response Surface Methodology analysis}\n")
        f.write("\\label{tab:summary}\n")
        f.write("\\begin{tabular}{lc}\n")
        f.write("\\toprule\n")
        f.write("Parameter & Value \\\\\n")
        f.write("\\midrule\n")
        
        for _, row in df_summary.iterrows():
            if row['Parameter'] == '':
                f.write("\\midrule\n")
            else:
                f.write(f"{row['Parameter']} & {row['Value']} \\\\\n")
        
        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\end{table}\n")
    
    print(f"✅ LaTeX summary table: {latex_summary}")

#%%============================================================================
# EXPORT TO EXCEL
#%%============================================================================

def export_to_excel(df_anova, df_effects, df_summary, stats_summary):
    """Export all tables to Excel with formatting"""
    
    print(f"\n{'='*50}")
    print("EXPORTING TO EXCEL")
    print(f"{'='*50}")
    
    excel_file = os.path.join(OUTPUT_FOLDER, "anova_tables_complete.xlsx")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: ANOVA Global
    ws1 = wb.active
    ws1.title = "ANOVA Global"
    
    ws1['A1'] = "Table 1: Global ANOVA"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:F1')
    
    # Headers
    for col, header in enumerate(['Source', 'SS', 'df', 'MS', 'F', 'p-value'], 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for r_idx, row in enumerate(dataframe_to_rows(df_anova, index=False, header=False), 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            if c_idx > 1:
                cell.alignment = Alignment(horizontal='right')
    
    # Note
    note_row = len(df_anova) + 5
    ws1.cell(row=note_row, column=1).value = f"Note: R² = {stats_summary['R2']:.4f}, Adjusted R² = {stats_summary['R2_adj']:.4f}, RMSE = {stats_summary['RMSE']:.4f}"
    ws1.cell(row=note_row, column=1).font = Font(italic=True, size=9)
    
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 12
    
    # Sheet 2: Individual Effects
    ws2 = wb.create_sheet("Individual Effects")
    
    ws2['A1'] = "Table 2: Significant Individual Effects"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')
    
    for col, header in enumerate(['Rank', 'Effect', 'Type', 'Coefficient', 'Std. Effect', 'p-value', 'Sig.'], 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for r_idx, row in enumerate(dataframe_to_rows(df_effects, index=False, header=False), 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            if c_idx in [4, 5, 6]:
                cell.alignment = Alignment(horizontal='right')
    
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 8
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    
    ws3['A1'] = "Table 3: Summary Statistics"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:B1')
    
    for col, header in enumerate(['Parameter', 'Value'], 1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=False), 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 25
    
    wb.save(excel_file)
    print(f"✅ Excel file: {excel_file}")

#%%============================================================================
# EXPORT TO CSV
#%%============================================================================

def export_to_csv(df_anova, df_effects, df_summary):
    """Export tables to CSV"""
    
    print(f"\n{'='*50}")
    print("EXPORTING TO CSV")
    print(f"{'='*50}")
    
    csv_folder = os.path.join(OUTPUT_FOLDER, "csv")
    os.makedirs(csv_folder, exist_ok=True)
    
    df_anova.to_csv(os.path.join(csv_folder, "table1_anova_global.csv"), index=False)
    df_effects.to_csv(os.path.join(csv_folder, "table2_individual_effects.csv"), index=False)
    df_summary.to_csv(os.path.join(csv_folder, "table3_summary.csv"), index=False)
    
    print(f"✅ CSV files saved in: {csv_folder}")

#%%============================================================================
# MAIN EXECUTION
#%%============================================================================

def main():
    """Main execution"""
    
    # Create output folder
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Load data
    all_effects, significant_effects, model_stats = load_analysis()
    
    # Calculate ANOVA
    anova_table, stats_summary = calculate_anova_global(all_effects, model_stats)
    
    # Create DataFrames
    df_anova = create_anova_global_df(anova_table)
    df_effects = create_individual_effects_df(significant_effects)
    df_summary = create_summary_df(all_effects, significant_effects, stats_summary)
    
    # Export to all formats
    export_to_latex(df_anova, df_effects, df_summary, stats_summary)
    export_to_excel(df_anova, df_effects, df_summary, stats_summary)
    export_to_csv(df_anova, df_effects, df_summary)
    
    # Display tables
    print(f"\n{'='*70}")
    print("TABLE 1: GLOBAL ANOVA")
    print(f"{'='*70}")
    print(df_anova.to_string(index=False))
    
    print(f"\n{'='*70}")
    print("TABLE 2: INDIVIDUAL SIGNIFICANT EFFECTS (Top {MAX_EFFECTS_INDIVIDUAL_TABLE})")
    print(f"{'='*70}")
    print(df_effects.to_string(index=False))
    
    print(f"\n{'='*70}")
    print("TABLE 3: SUMMARY STATISTICS")
    print(f"{'='*70}")
    print(df_summary.to_string(index=False))
    
    # Final summary
    print(f"\n{'='*70}")
    print("🎉 COMPLETE ANOVA TABLES GENERATED")
    print(f"{'='*70}")
    
    print(f"\n📊 Tables Generated:")
    print(f"\n   LaTeX format:")
    print(f"   - table1_anova_global.tex")
    print(f"   - table2_individual_effects.tex")
    print(f"   - table3_summary.tex")
    
    print(f"\n   Excel format:")
    print(f"   - anova_tables_complete.xlsx (3 sheets)")
    
    print(f"\n   CSV format:")
    print(f"   - table1_anova_global.csv")
    print(f"   - table2_individual_effects.csv")
    print(f"   - table3_summary.csv")
    
    print(f"\n📁 Output location: {OUTPUT_FOLDER}/")
    
    print(f"\n💡 For your paper:")
    print(f"   Table 1: Global ANOVA (validates model)")
    print(f"   Table 2: Individual effects (identifies important variables)")
    print(f"   Table 3: Summary (for methodology section)")
    
    print(f"\n{'='*70}")

if __name__ == "__main__":
    main()
